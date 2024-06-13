import time
import pyodbc
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

#cria a conecxão ao banco
def get_db_connection():
    try:
        conn = pyodbc.connect(
            'DRIVER={SQL Server};'
            'SERVER=Server Name;'
            'DATABASE=DataBase Name;'
            'UID=User Name;'
            'PWD=Password;'
            'Trusted_Connection=no;'
        )
        return conn
    #checa a conexão e retorna em caso de erro
    except pyodbc.Error as e:
        print("Erro na conexão:", e)
        return None
    
#conecta ao banco
def fetch_data():
    conn = get_db_connection()
    #checa e retorna caso a conexão falhe
    if conn is None:
        print("Falha na conexão a banco de dados.")
        return []
    #executa o SQL
    try:
        cur = conn.cursor()
        cur.execute('''
         SELECT 
	    LOCALIZACAO, 
	    DATA,
	    CLI_NOME,
	    CLI_CPF,
	    UF,
	    CIDADE,
	    VENDEDOR,
	    PRODUTO,
	    QUANTIDADE, 
	    VALOR_TOTAL
	FROM VW_FAT_DET
    JOIN CLIENTE C ON CLI_COD = CD_CLI_GER
    WHERE FOR_CODI = '01076'
    AND CONVERT(DATETIME, DATA, 103)>= DATEADD(MONTH, DATEDIFF(MONTH, 0, GETDATE()) - 1, 0)
    AND CONVERT(DATETIME, DATA, 103) < DATEADD(MONTH, DATEDIFF(MONTH, 0, GETDATE()), 0)
    AND NATUREZA IN ('VEN')
    ORDER BY DATA ASC
        ''')
        #retorna os dados recolhidos no select
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return rows
    #retorna em caso de erro no recolhimento dos dados
    except pyodbc.Error as e:
        print("Erro ao juntar os dados data:", e)
        return []

def save_to_excel(rows, filename='extraido/RELATORIO.xlsx'):
    # Verifica se o diretório existe, caso contrário, cria
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    # Adiciona uma verificação para o formato dos dados retornados
    #if len(rows) > 0:
        # Printa a primeira linha da estrutura de dados
        #print(f"First row: {rows[0]}")
        # Checa o tamanho da primeira linha se fecha com a quantidade de colunas solicitadas 
        #if len(rows[0]) != 10:
            #print(f"Unexpected data shape: {len(rows[0])} columns instead of 10")
            #return

    # Transforma as linhas em uma lista de tuplas
    data = [tuple(row) for row in rows]
    #printa as informações recolhidas e transformadas
    ##print(f"Data rows: {data}")

    # Cria um DataFrame
    df = pd.DataFrame(data, columns=
    ['DATA FATURAMENTO','RAZÃO SOCIAL','VENDEDOR','ITEM','QTDE','VALOR'])
    df.to_excel(filename, index=False)
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Definir o fundo vermelho para a primeira linha
    red_fill = PatternFill(start_color="FFFF5050", end_color="FFFF5050", fill_type="solid")
    white_font = Font(color="FFFFFF", name="Aptos Narrow", bold=True)
    for cell in sheet[1]:
        cell.fill = red_fill
        cell.font = white_font

    # Salvar o arquivo modificado
    workbook.save(filename)
    print(f"Data saved to {filename}")
# checa se os dados retornam vazios
if __name__ == '__main__':
    data = fetch_data()
    #caso retorne informação salva em arquivo xls
    if data:
        save_to_excel(data)
    # em caso de dado vazio retorna o erro
    
    else:
        print("Sem informações, erro ao recolher os dados.")

#Checa a conexão com o servidor smtp e retorna o status
def check_smtp_connection(smtp_server, smtp_port, smtp_user, smtp_password):
    try:
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(smtp_user, smtp_password)
        print("Conexão com o servidor SMTP bem-sucedida.")
        return True
    except Exception as e:
        print(f"Falha na conexão com o servidor SMTP: {e}")
        return False

def send_email_with_attachment(smtp_server, smtp_port, smtp_user, smtp_password, from_addr, to_addr, subject, body, file_path):
    #verifica a conexão e em caso de erro tenta novamente em 30 minutos
    while not check_smtp_connection(smtp_server, smtp_port, smtp_user, smtp_password):
        print("Tentando novamente em 30 minutos...")
        time.sleep(1800)  # Espera por 30 minutos
        
   # Cria a mensagem
    msg = MIMEMultipart()
    msg['From'] = from_addr
    msg['To'] = to_addr
    msg['Subject'] = subject

    # Adiciona o corpo do email
    msg.attach(MIMEText(body, 'plain'))

    # Anexa o arquivo
    with open(file_path, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename= {os.path.basename(file_path)}')
        msg.attach(part)

    # Configura o servidor SMTP e envia o email
    with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
        server.login(smtp_user, smtp_password)
        server.sendmail(from_addr, to_addr, msg.as_string())

    print("Email enviado com sucesso!")

# Dados do email
smtp_server = 'host smtp'
smtp_port = 'porta smtp' 
smtp_user = 'email remetente host'
smtp_password = 'senha email remetente'
from_addr = 'email remetente '
to_addr = 'email receptor'  
subject = 'assunto'
body = 'corpo do email'
file_path = 'local do anexo/arquivo'

# Envia o email
send_email_with_attachment(smtp_server, smtp_port, smtp_user, smtp_password, from_addr, to_addr, subject, body, file_path)
